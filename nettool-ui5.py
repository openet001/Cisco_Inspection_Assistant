#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from openpyxl import load_workbook
from netmiko import ConnectHandler
import requests
import openai

# ======= 通用API模型接口配置 =======
# 若模型调用出错，请参考官网API-Key的调用说明对代码进行修改
AI_MODEL_CONFIG = {
    "Deepseek": {
        "api_key_label": "Deepseek API-Key",  # 申请key并实名认证，api调用需要充值，网页对话不需要充值
        "api_url": "https://api.deepseek.com/v1/chat/completions",
        "header_key": "Authorization",
        "header_prefix": "Bearer",
        "models": ["deepseek-chat", "deepseek-reasoner"]
    },
    "腾讯元宝": {
        "api_key_label": "腾讯元宝 API-Key",
        "api_url": "https://models.tencent.com/v1/chat/completions",  # 请按实际填写
        "header_key": "Authorization",
        "header_prefix": "Bearer",
        "models": ["tencent-ernie-bot", "tencent-ernie-turbo"]
    },
    "通义千问": {
        "api_key_label": "通义千问 API-Key",
        "api_url": "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation",  # 请按实际填写
        "header_key": "Authorization",
        "header_prefix": "Bearer",
        "models": ["qwen-turbo", "qwen-plus"]
    },
    "火山引擎": {
        "api_key_label": "火山引擎 API-Key",
        "api_url": "https://api.volcengine.com/llm/v1/chat/completions",  # 请按实际填写
        "header_key": "Authorization",
        "header_prefix": "Bearer",
        "models": ["skylark-chat", "skylark2-lite"]
    },
    "ChatGPT": {
        "api_key_label": "OpenAI API-Key",
        "api_url": "https://api.openai.com/v1/chat/completions",
        "header_key": "Authorization",
        "header_prefix": "Bearer",
        "models": ["gpt-3.5-turbo", "gpt-4o", "gpt-4-turbo"]
    }
}

def universal_ai_analyze(model_type, api_key, model_name, log_text):
    if not api_key:
        return "未设置API Key，无法分析。"
    if model_type == "Deepseek":
        # 用 openai SDK 调用 Deepseek
        try:
            client = openai.OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "你是一名网络安全运维专家，请对以下巡检日志做智能分析，包括异常、告警、建议。"},
                    {"role": "user", "content": log_text}
                ],
                stream=False
            )
            return response.choices[0].message.content
        except Exception as e:
            return f"分析失败: {e}"
    else:
        # 其它模型继续用 requests 方式
        config = AI_MODEL_CONFIG.get(model_type)
        if not config:
            return "未支持的模型类型。"
        api_url = config["api_url"]
        headers = {
            config["header_key"]: f"{config['header_prefix']} {api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "你是一名网络安全运维专家，请对以下巡检日志做智能分析，包括异常、告警、建议。"},
                {"role": "user", "content": log_text}
            ],
            "temperature": 0.2
        }
        try:
            resp = requests.post(api_url, json=data, headers=headers, timeout=60)
            resp.raise_for_status()
            result = resp.json()
            if "choices" in result and result["choices"]:
                return result['choices'][0]['message'].get('content', "") or str(result)
            elif "output" in result:
                return result["output"] or str(result)
            elif "result" in result:
                return result["result"]
            else:
                return str(result)
        except Exception as e:
            return f"分析失败: {e}"

class DeviceInspectorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("网络设备巡检助手")
        self.root.geometry("950x780")

        # 设备缓存
        self.device_cache = {}
        self.device_ip_list = tk.StringVar(value=[])

        # 当前设备信息
        self.template_file = tk.StringVar()
        self.selected_ip = tk.StringVar()
        self.protocol_var = tk.StringVar()
        self.device_type_var = tk.StringVar()
        self.cmds_var = tk.StringVar()
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.secret_var = tk.StringVar()
        self.port_var = tk.StringVar()
        self.log_text = ""

        # AI模型相关
        self.ai_model_var = tk.StringVar(value="Deepseek")
        self.ai_model_name_var = tk.StringVar()
        self.api_key_var = tk.StringVar()
        self.api_key_entry = None
        self.api_key_hidden = True

        self._build_gui()

    def _build_gui(self):
        frm = ttk.Frame(self.root)
        frm.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(frm, text="智能分析模型:").grid(row=0, column=0, sticky='e')
        self.ai_model_box = ttk.Combobox(frm, textvariable=self.ai_model_var, values=list(AI_MODEL_CONFIG.keys()), width=14, state="readonly")
        self.ai_model_box.grid(row=0, column=1, sticky='w')
        self.ai_model_box.bind("<<ComboboxSelected>>", self.on_ai_model_changed)

        ttk.Label(frm, text="模型名称:").grid(row=0, column=2, sticky='e')
        self.ai_model_name_box = ttk.Combobox(frm, textvariable=self.ai_model_name_var, width=18)
        self.ai_model_name_box.grid(row=0, column=3, sticky='w')

        ttk.Label(frm, text="API-Key:").grid(row=1, column=0, sticky='e')
        self.api_key_entry = ttk.Entry(frm, textvariable=self.api_key_var, width=50, show="*")
        self.api_key_entry.grid(row=1, column=1, columnspan=2, sticky='w')
        ttk.Button(frm, text="显示/隐藏", command=self.toggle_api_key).grid(row=1, column=3, padx=8, sticky='w')

        ttk.Label(frm, text="Excel模板:").grid(row=2, column=0, sticky='e')
        ttk.Entry(frm, textvariable=self.template_file, width=50).grid(row=2, column=1)
        ttk.Button(frm, text="选择", command=self.select_template).grid(row=2, column=2, padx=5)
        ttk.Button(frm, text="加载模板设备", command=self.load_devices_from_excel).grid(row=2, column=3, padx=8)

        ttk.Label(frm, text="设备IP:").grid(row=3, column=0, sticky='e')
        self.ip_combobox = ttk.Combobox(frm, textvariable=self.selected_ip, values=[], width=18, postcommand=self.update_ip_combobox)
        self.ip_combobox.grid(row=3, column=1, sticky='w')
        self.ip_combobox.bind("<<ComboboxSelected>>", self.on_ip_selected)

        ttk.Label(frm, text="登录方式:").grid(row=3, column=2, sticky='e')
        self.protocol_box = ttk.Combobox(frm, textvariable=self.protocol_var, values=["ssh", "telnet"], width=8)
        self.protocol_box.grid(row=3, column=3, sticky='w')

        ttk.Label(frm, text="设备类型:").grid(row=4, column=0, sticky='e')
        self.dtype_box = ttk.Combobox(frm, textvariable=self.device_type_var,
                     values=["cisco_ios", "huawei", "ruijie", "cisco_xe", "cisco_nxos", "h3c", "other"], width=12)
        self.dtype_box.grid(row=4, column=1, sticky='w')

        ttk.Label(frm, text="端口:").grid(row=4, column=2, sticky='e')
        ttk.Entry(frm, textvariable=self.port_var, width=8).grid(row=4, column=3, sticky='w')

        ttk.Label(frm, text="用户名:").grid(row=5, column=0, sticky='e')
        ttk.Entry(frm, textvariable=self.username_var, width=16).grid(row=5, column=1, sticky='w')
        ttk.Label(frm, text="密码:").grid(row=5, column=2, sticky='e')
        ttk.Entry(frm, textvariable=self.password_var, show='*', width=16).grid(row=5, column=3, sticky='w')
        ttk.Label(frm, text="Enable密码:").grid(row=6, column=0, sticky='e')
        ttk.Entry(frm, textvariable=self.secret_var, show='*', width=16).grid(row=6, column=1, sticky='w')

        ttk.Label(frm, text="巡检命令(逗号分隔):").grid(row=7, column=0, sticky='e')
        ttk.Entry(frm, textvariable=self.cmds_var, width=60).grid(row=7, column=1, columnspan=3, sticky='w')

        ttk.Button(frm, text="连接测试", command=self.test_connect).grid(row=8, column=0, pady=8)
        ttk.Button(frm, text="执行巡检", command=self.start_inspect_thread).grid(row=8, column=1, pady=8)
        ttk.Button(frm, text="分析日志", command=self.start_analysis_thread).grid(row=8, column=2, pady=8)
        ttk.Button(frm, text="清空输出", command=self.clear_output).grid(row=8, column=3, pady=8)

        self.output_box = scrolledtext.ScrolledText(self.root, width=110, height=25, font=("Consolas", 11))
        self.output_box.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        self.on_ai_model_changed()  # 初始化模型下拉

    def toggle_api_key(self):
        if self.api_key_hidden:
            self.api_key_entry.config(show="")
            self.api_key_hidden = False
        else:
            self.api_key_entry.config(show="*")
            self.api_key_hidden = True

    def on_ai_model_changed(self, event=None):
        model = self.ai_model_var.get()
        names = AI_MODEL_CONFIG.get(model, {}).get("models", [])
        if names:
            self.ai_model_name_box["values"] = names
            self.ai_model_name_var.set(names[0])
        else:
            self.ai_model_name_box["values"] = []
            self.ai_model_name_var.set("")
        # 更新key label
        label = AI_MODEL_CONFIG.get(model, {}).get("api_key_label", "API-Key")
        self.api_key_entry.master.grid_slaves(row=1, column=0)[0].config(text=label)

    def select_template(self):
        f = filedialog.askopenfilename(title="选择Excel模板", filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if f:
            self.template_file.set(f)

    def load_devices_from_excel(self):
        path = self.template_file.get()
        if not os.path.exists(path):
            messagebox.showerror("错误", f"模板文件不存在: {path}")
            return
        try:
            wb = load_workbook(path)
            ws = wb[wb.sheetnames[0]]
            self.device_cache.clear()
            ip_list = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_col=9), 2):
                if str(row[1].value).strip() == '#':
                    continue
                ip = str(row[2].value).strip()
                protocol = str(row[3].value).strip().lower() if row[3].value else "ssh"
                port = int(row[4].value) if row[4].value else (22 if protocol == "ssh" else 23)
                username = str(row[5].value).strip() if row[5].value else ""
                password = str(row[6].value).strip() if row[6].value else ""
                secret = str(row[7].value).strip() if row[7].value else ""
                device_type = str(row[8].value).strip() if row[8].value else "cisco_ios"
                # 读取命令
                cmd_list = []
                if row[8].value and row[8].value in wb.sheetnames:
                    cmd_sheet = wb[row[8].value]
                    for cmdrow in cmd_sheet.iter_rows(min_row=2, max_col=2):
                        if str(cmdrow[0].value).strip() != "#" and cmdrow[1].value:
                            cmd_list.append(str(cmdrow[1].value).strip())
                devinfo = {
                    "ip": ip,
                    "host": ip,
                    "protocol": protocol,
                    "port": port,
                    "username": username,
                    "password": password,
                    "secret": secret,
                    "device_type": device_type,
                    "cmd_list": cmd_list
                }
                self.device_cache[ip] = devinfo
                ip_list.append(ip)
            self.device_ip_list.set(ip_list)
            self.ip_combobox['values'] = ip_list
            if ip_list:
                self.selected_ip.set(ip_list[0])
                self.fill_device_fields(ip_list[0])
            self.output_box.insert(tk.END, f"模板设备加载完成，共{len(ip_list)}台。\n\n")
            wb.close()
        except Exception as e:
            messagebox.showerror("Excel加载失败", str(e))

    def update_ip_combobox(self):
        self.ip_combobox['values'] = list(self.device_cache.keys())

    def on_ip_selected(self, event=None):
        ip = self.selected_ip.get()
        self.fill_device_fields(ip)

    def fill_device_fields(self, ip):
        dev = self.device_cache.get(ip)
        if not dev:
            return
        self.protocol_var.set(dev.get("protocol", "ssh"))
        self.device_type_var.set(dev.get("device_type", "cisco_ios"))
        self.port_var.set(str(dev.get("port", 22)))
        self.username_var.set(dev.get("username", ""))
        self.password_var.set(dev.get("password", ""))
        self.secret_var.set(dev.get("secret", ""))
        self.cmds_var.set(",".join(dev.get("cmd_list", [])))

    def test_connect(self):
        device = self._get_device_info()
        if not device:
            return
        def worker():
            try:
                conn = ConnectHandler(**device, conn_timeout=8)
                prompt = conn.find_prompt()
                self.output_box.insert(tk.END, f"[{device['ip']}] 连接成功, 提示符: {prompt}\n")
                conn.disconnect()
            except Exception as e:
                self.output_box.insert(tk.END, f"[{device['ip']}] 连接失败: {e}\n")
        threading.Thread(target=worker, daemon=True).start()

    def start_inspect_thread(self):
        threading.Thread(target=self.inspect_device, daemon=True).start()

    def inspect_device(self):
        device = self._get_device_info()
        if not device:
            return
        cmds = [x.strip() for x in self.cmds_var.get().split(',') if x.strip()]
        if not cmds:
            self.output_box.insert(tk.END, "[提示] 巡检命令不能为空！\n")
            return
        try:
            self.output_box.insert(tk.END, f"开始连接设备 {device['ip']} ...\n")
            conn = ConnectHandler(**device, conn_timeout=12)
            if self.secret_var.get():
                try:
                    conn.enable()
                except Exception:
                    pass
            hostname = conn.find_prompt()
            log_lines = [f"设备: {device['ip']} ({hostname})"]
            for cmd in cmds:
                self.output_box.insert(tk.END, f"执行: {cmd}\n")
                try:
                    output = conn.send_command(cmd, read_timeout=18)
                    log_lines.append(f"\n====== {cmd} ======\n{output}\n")
                except Exception as e:
                    log_lines.append(f"\n====== {cmd} 执行失败: {e} ======\n")
            conn.disconnect()
            # 保存日志
            log_dir = "inspection_logs"
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(log_dir, f"{device['ip']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
            log_content = '\n'.join(log_lines)
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(log_content)
            self.log_text = log_content
            self.output_box.insert(tk.END, f"\n巡检完成，日志已保存至: {log_file}\n")
        except Exception as e:
            self.output_box.insert(tk.END, f"巡检失败: {e}\n")

    def start_analysis_thread(self):
        threading.Thread(target=self.analyze_log, daemon=True).start()

    def analyze_log(self):
        if not self.log_text:
            self.output_box.insert(tk.END, "[提示] 请先完成巡检以生成日志！\n")
            return
        model_type = self.ai_model_var.get()
        api_key = self.api_key_var.get().strip()
        model_name = self.ai_model_name_var.get().strip()
        self.output_box.insert(tk.END, f"\n[{model_type}] 正在智能分析日志，请稍候...\n")
        result = universal_ai_analyze(model_type, api_key, model_name, self.log_text[:3000])
        self.output_box.insert(tk.END, f"\n【{model_type}分析结果】\n{result}\n\n")

    def clear_output(self):
        self.output_box.delete(1.0, tk.END)

    def _get_device_info(self):
        ip = self.selected_ip.get().strip()
        if not ip:
            messagebox.showwarning("必填", "请选择IP地址")
            return None
        prot = self.protocol_var.get().strip().lower()
        dtype = self.device_type_var.get().strip()
        port = self.port_var.get().strip()
        user = self.username_var.get().strip()
        pwd = self.password_var.get().strip()
        secret = self.secret_var.get().strip()
        device = {
            "device_type": dtype,
            "host": ip,
            "ip": ip,
            "username": user,
            "password": pwd,
            "port": int(port) if port else (22 if prot == "ssh" else 23)
        }
        if secret:
            device["secret"] = secret
        if prot == "telnet":
            device["device_type"] = dtype + "_telnet"
        return device

if __name__ == "__main__":
    root = tk.Tk()
    app = DeviceInspectorApp(root)
    root.mainloop()